"""
Exportación a Excel.

Genera un libro con el mismo espíritu que los "Control de Obra" del usuario:
portada con ficha y cuadro de mando, una hoja por módulo, cabeceras congeladas,
autofiltro, anchos calculados y formato condicional en los estados.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .. import db
from ..schema import ENTITIES, MENU, Entity
from . import kpis

AZUL = "1E3A5F"
AZUL_CLARO = "DCE6F1"
GRIS = "F2F2F2"
VERDE = "C6EFCE"
AMBAR = "FFEB9C"
ROJO = "FFC7CE"

_BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)

COLOR_ESTADO = {
    # tareas
    "Completada": VERDE, "En curso": AMBAR, "Bloqueada": ROJO, "Cancelada": GRIS,
    # dispositivos
    "Instalado": VERDE, "Probado": VERDE, "Entregado": VERDE,
    "Configurado": AMBAR, "Conexionado": AMBAR, "Recibido en obra": AMBAR,
    "Averiado": ROJO, "Pendiente sustitución": ROJO,
    # incidencias
    "Abierta": ROJO, "Escalada": ROJO, "En gestión": AMBAR,
    "Resuelta": VERDE, "Cerrada": VERDE,
    # personal / documentos
    "ACCESS": VERDE, "NO ACCESS": ROJO, "VALIDANDO": AMBAR, "BAJA": GRIS,
    "Pendiente": AMBAR, "Solicitada": AMBAR, "Rechazada": ROJO, "Caducada": ROJO,
    "OK": VERDE, "REPONER": AMBAR, "EXCEDIDO": ROJO, "AGOTADO": ROJO,
}


def _cabecera(ws: Worksheet, titulo: str, columnas: list[str]) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columnas), 1))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    for i, nombre in enumerate(columnas, start=1):
        c = ws.cell(row=2, column=i, value=nombre)
        c.font = Font(bold=True, color=AZUL)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        c.border = _BORDE
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"


def _autoajustar(ws: Worksheet, columnas: list[str], filas: list[list]) -> None:
    for i, nombre in enumerate(columnas, start=1):
        ancho = len(str(nombre)) + 4
        for fila in filas[:300]:
            v = fila[i - 1] if i - 1 < len(fila) else ""
            ancho = max(ancho, min(len(str(v or "")) + 3, 48))
        ws.column_dimensions[get_column_letter(i)].width = min(ancho, 50)


def _valor_celda(f, valor, refs: dict):
    if valor is None:
        return ""
    if f.type == "ref":
        return refs.get((f.ref, int(valor)), valor)
    if f.type == "bool":
        return "Sí" if valor else "No"
    if f.type == "date":
        try:
            return datetime.fromisoformat(str(valor)[:10]).date()
        except ValueError:
            return valor
    return valor


def _refs_de(ent: Entity) -> dict:
    """Mapa (entidad, id) -> etiqueta, para pintar las referencias legibles."""
    cache: dict[tuple[str, int], str] = {}
    for f in ent.fields:
        if f.type == "ref" and f.ref:
            destino = ENTITIES[f.ref]
            for r in db.query(f"SELECT id, {destino.title_field} t FROM {f.ref}"):
                cache[(f.ref, r["id"])] = r["t"]
    return cache


def hoja_entidad(wb: Workbook, key: str, obra_id: int | None) -> None:
    ent = ENTITIES[key]
    filas_db = db.listar(key, obra_id if ent.per_obra else None)
    campos = [f for f in ent.fields]
    columnas = [f.label for f in campos]
    refs = _refs_de(ent)

    ws = wb.create_sheet(ent.plural[:31])
    _cabecera(ws, f"{ent.plural.upper()}", columnas)

    datos = []
    for fila in filas_db:
        datos.append([_valor_celda(f, fila.get(f.name), refs) for f in campos])

    idx_estado = next((i for i, f in enumerate(campos) if f.name == "estado"), None)
    for r, valores in enumerate(datos, start=3):
        for cidx, v in enumerate(valores, start=1):
            c = ws.cell(row=r, column=cidx, value=v)
            c.border = _BORDE
            c.alignment = Alignment(vertical="top", wrap_text=len(str(v or "")) > 40, indent=1)
            f = campos[cidx - 1]
            if f.type == "money":
                c.number_format = '#,##0.00 "€"'
            elif f.type == "percent":
                c.number_format = "0.0"
            elif f.type == "date":
                c.number_format = "DD/MM/YYYY"
        if idx_estado is not None:
            color = COLOR_ESTADO.get(str(valores[idx_estado]))
            if color:
                ws.cell(row=r, column=idx_estado + 1).fill = PatternFill("solid", fgColor=color)

    if datos:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(columnas))}{len(datos) + 2}"
    _autoajustar(ws, columnas, datos)


def hoja_portada(wb: Workbook, obra_id: int) -> None:
    r = kpis.resumen(obra_id)
    obra, kpi, eco = r["obra"], r["tareas"], r["economico"]

    ws = wb.create_sheet("Portada", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 22

    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "CONTROL INTEGRAL DE OBRA"
    c.font = Font(bold=True, size=20, color=AZUL)
    ws.merge_cells("B3:F3")
    ws["B3"] = "Seguridad · CCTV · Intrusión · Control de accesos · Redes"
    ws["B3"].font = Font(italic=True, size=10, color="666666")

    def bloque(fila: int, titulo: str, pares: list[tuple[str, object]], col: str = "B"):
        col2 = chr(ord(col) + 1)
        ws.merge_cells(f"{col}{fila}:{col2}{fila}")
        c = ws[f"{col}{fila}"]
        c.value = titulo
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(indent=1, vertical="center")
        for i, (k, v) in enumerate(pares, start=1):
            a = ws[f"{col}{fila + i}"]
            a.value = k
            a.font = Font(bold=True, size=10)
            a.fill = PatternFill("solid", fgColor=GRIS)
            a.alignment = Alignment(indent=1)
            b = ws[f"{col2}{fila + i}"]
            b.value = v
            b.alignment = Alignment(indent=1)
        return fila + len(pares) + 2

    f = bloque(5, "FICHA DEL PROYECTO", [
        ("Proyecto", obra.get("nombre")),
        ("Código / OT", obra.get("codigo")),
        ("Cliente", obra.get("cliente")),
        ("Cliente final", obra.get("cliente_final")),
        ("Emplazamiento", f"{obra.get('poblacion') or ''} {obra.get('provincia') or ''}".strip()),
        ("Dirección", obra.get("direccion")),
        ("Jefe de obra", obra.get("jefe_obra")),
        ("Interlocutor cliente", obra.get("interlocutor_cliente")),
        ("Estado", obra.get("estado")),
        ("Inicio", obra.get("fecha_inicio")),
        ("Fin previsto", obra.get("fecha_fin_prevista")),
    ])

    bloque(5, "CUADRO DE MANDO", [
        ("Avance global", f"{kpi['avance']} %"),
        ("Tareas totales", kpi["total"]),
        ("Completadas", kpi["completadas"]),
        ("Con retraso", kpi["retrasadas"]),
        ("Bloqueadas", kpi["bloqueadas"]),
        ("Dispositivos", r["dispositivos"]["total"]),
        ("Instalados", r["dispositivos"]["instalados"]),
        ("Incidencias abiertas", r["otros"]["incidencias_abiertas"]),
        ("Documentos pendientes", r["otros"]["docs_pendientes"]),
        ("Personal con acceso", r["otros"]["personal_con_acceso"]),
    ], col="E")

    f = bloque(f, "ECONÓMICO", [
        ("Importe de contrato", eco["contrato"]),
        ("Ampliaciones aprobadas", eco["ampliaciones"]),
        ("Ingreso total", eco["ingresos"]),
        ("Coste incurrido", eco["coste_total"]),
        ("Margen bruto", eco["margen"]),
        ("Margen %", f"{eco['margen_pct']} %"),
        ("Certificado", eco["certificado"]),
        ("Pendiente de cobro", eco["pendiente_cobro"]),
    ])
    for fila in range(6, f):
        celda = ws[f"C{fila}"]
        if isinstance(celda.value, (int, float)) and not isinstance(celda.value, bool):
            celda.number_format = '#,##0.00 "€"'

    if r["alertas"]:
        ws[f"B{f}"] = "PUNTOS DE ATENCIÓN"
        ws[f"B{f}"].font = Font(bold=True, color="FFFFFF")
        ws[f"B{f}"].fill = PatternFill("solid", fgColor="C0392B")
        ws.merge_cells(f"B{f}:F{f}")
        for i, a in enumerate(r["alertas"][:14], start=1):
            ws[f"B{f + i}"] = a["severidad"].upper()
            ws[f"B{f + i}"].font = Font(bold=True, size=9)
            ws.merge_cells(f"C{f + i}:F{f + i}")
            ws[f"C{f + i}"] = f"{a['titulo']} — {a['detalle']}"
            ws[f"C{f + i}"].alignment = Alignment(wrap_text=True, vertical="top")
            color = {"critica": ROJO, "alta": AMBAR, "media": GRIS}.get(a["severidad"])
            if color:
                ws[f"B{f + i}"].fill = PatternFill("solid", fgColor=color)

    ws[f"B{f + 18}"] = f"Generado por ObraSec · {datetime.now():%d/%m/%Y %H:%M}"
    ws[f"B{f + 18}"].font = Font(italic=True, size=9, color="888888")


def hoja_stock(wb: Workbook, obra_id: int) -> None:
    filas = kpis.stock(obra_id)
    if not filas:
        return
    columnas = ["Material", "Categoría", "Ud.", "Proveedor", "Recibido", "Gastado",
                "Restante", "% Consumido", "Stock mín.", "Precio ud.",
                "Valor gastado", "Valor restante", "Alerta"]
    ws = wb.create_sheet("Stock")
    _cabecera(ws, "STOCK DE MATERIAL · recibido menos consumido", columnas)
    datos = []
    for s in filas:
        datos.append([
            s.get("material"), s.get("categoria"), s.get("unidad"), s.get("proveedor"),
            s.get("recibido"), s.get("gastado"), s.get("restante"), s.get("pct_consumido"),
            s.get("stock_min"), s.get("precio"), s.get("valor_gastado"),
            s.get("valor_restante"), s.get("alerta"),
        ])
    for r, valores in enumerate(datos, start=3):
        for cidx, v in enumerate(valores, start=1):
            c = ws.cell(row=r, column=cidx, value=v)
            c.border = _BORDE
            if cidx in (10, 11, 12):
                c.number_format = '#,##0.00 "€"'
        color = COLOR_ESTADO.get(str(valores[-1]))
        if color:
            ws.cell(row=r, column=len(columnas)).fill = PatternFill("solid", fgColor=color)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(columnas))}{len(datos) + 2}"
    _autoajustar(ws, columnas, datos)


def libro_obra(obra_id: int) -> bytes:
    """Libro completo de la obra: portada + todas las hojas."""
    wb = Workbook()
    wb.remove(wb.active)
    hoja_portada(wb, obra_id)
    for key in MENU:
        if key in ENTITIES:
            hoja_entidad(wb, key, obra_id)
    hoja_stock(wb, obra_id)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def libro_entidad(key: str, obra_id: int | None) -> bytes:
    """Una sola hoja: por ejemplo el inventario de instalación."""
    wb = Workbook()
    wb.remove(wb.active)
    hoja_entidad(wb, key, obra_id)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nombre_archivo(obra_id: int | None, sufijo: str) -> str:
    obra = db.obtener("obras", obra_id) if obra_id else None
    base = (obra or {}).get("codigo") or (obra or {}).get("nombre") or "ObraSec"
    limpio = "".join(ch for ch in str(base) if ch.isalnum() or ch in " -_")[:40].strip()
    return f"{limpio}_{sufijo}_{date.today():%Y%m%d}"
