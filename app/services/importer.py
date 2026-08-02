"""
Importación desde los libros Excel de control de obra.

Funciona con los tres formatos que ya usa el usuario ("Control de Obra PRO",
"Control_de_Obra_Sistema_Integral" y la hoja original de Google Sheets) y con
cualquier libro parecido, porque el emparejamiento de columnas es por
similitud del texto de la cabecera, no por posición fija.

Estrategia:
  1. Localizar la fila de cabecera de cada hoja (la que tiene más etiquetas
     cortas de texto en las 8 primeras filas).
  2. Emparejar cada cabecera con un campo de la entidad destino usando el
     nombre del campo, su etiqueta y una tabla de sinónimos.
  3. Insertar las filas con datos, ignorando las de ejemplo y las vacías.
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from .. import db
from ..schema import ENTITIES

# Nombre de hoja (normalizado) -> entidad destino
HOJAS = {
    "tareas": "tareas", "hitos": "tareas", "planificacion": "tareas",
    "dispositivos": "dispositivos", "equipos": "dispositivos", "inventario": "dispositivos",
    "materiales": "materiales", "material": "materiales", "almacen": "materiales",
    "consumos": "consumos", "consumo": "consumos",
    "personal": "personal", "mano de obra": "personal",
    "subcontratas": "subcontratas", "subcontratistas": "subcontratas",
    "maquinaria": "maquinaria", "medios auxiliares": "maquinaria",
    "incidencias": "incidencias",
    "documentos": "documentos", "documentacion": "documentos",
    "planos": "planos",
    "ofertas": "ofertas", "ampliaciones": "ofertas",
    "presupuesto": "partidas", "partidas": "partidas",
    "certificaciones": "certificaciones",
    "visitas": "visitas",
    "contactos": "contactos", "agenda": "contactos",
    "partes": "partes",
}

# Sinónimos de cabecera -> nombre de campo. Se comparan normalizados.
SINONIMOS = {
    # tareas
    "tarea": "tarea", "descripcion de la tarea": "tarea", "actividad": "tarea",
    "categoria": "categoria", "tipo": "tipo",
    "prioridad": "prioridad", "estado": "estado",
    "avance": "avance", "% avance": "avance", "porcentaje": "avance",
    "fecha de inicio": "fecha_inicio", "inicio": "fecha_inicio",
    "fecha inicio": "fecha_inicio",
    "fecha de fin prevista": "fecha_fin", "fin previsto": "fecha_fin",
    "fecha fin prevista": "fecha_fin", "fin": "fecha_fin",
    "fecha de finalizacion": "fecha_fin_real", "fin real": "fecha_fin_real",
    "fecha fin real": "fecha_fin_real",
    "responsable": "responsable", "responsable / subcontrata": "responsable",
    "subcontrata / responsable": "responsable",
    "personal": "equipo", "personal asignado": "equipo",
    "coste previsto": "coste_previsto", "coste estimado": "coste_previsto",
    "coste estimado (e)": "coste_previsto",
    "notas": "notas", "observaciones": "notas",
    "tiempo ( d ; h )": "duracion_h", "duracion est. (dias)": "duracion_h",
    # dispositivos
    "codigo": "codigo", "etiqueta": "etiqueta", "nombre": "nombre",
    "tipo de dispositivo": "tipo", "marca": "marca", "modelo": "modelo",
    "n de serie": "num_serie", "n serie": "num_serie", "numero de serie": "num_serie",
    "serie": "num_serie",
    "ubicacion / zona": "zona", "ubicacion": "ubicacion", "zona": "zona",
    "cantidad": "cantidad", "cant.": "cantidad",
    "direccion ip / config": "ip", "ip / config": "ip", "direccion ip": "ip",
    "ip": "ip", "mac": "mac", "direccion mac": "mac",
    "fecha instalacion": "fecha_instalacion", "instalado": "fecha_instalacion",
    "fecha de instalacion": "fecha_instalacion",
    "responsable instalacion": "tecnico", "tecnico": "tecnico",
    "garantia hasta": "garantia_hasta",
    # materiales
    "material": "material", "unidad": "unidad", "ud.": "unidad",
    "cantidad recibida": "recibido", "recibido": "recibido",
    "recibida dia": "fecha_recepcion",
    "fecha ultima recepcion": "fecha_recepcion", "fecha recepcion": "fecha_recepcion",
    "proveedor": "proveedor",
    "precio unitario": "precio", "precio unitario (e)": "precio", "precio ud.": "precio",
    "precio": "precio", "precio hora": "precio_hora", "precio hora (e)": "precio_hora",
    "stock minimo": "stock_min", "stock min.": "stock_min",
    # consumos
    "fecha": "fecha", "codigo tarea": "_tarea_codigo", "cod. tarea": "_tarea_codigo",
    "registrado por": "registrado_por",
    # personal
    "nombre y apellidos": "nombre", "empresa": "empresa",
    "empresa / subcontrata": "empresa", "dni": "dni", "telefono": "telefono",
    "oficio": "oficio", "oficio / categoria": "oficio",
    "horas": "horas", "horas totales": "horas",
    "fecha incorporacion": "fecha_alta", "alta": "fecha_alta",
    "email": "email",
    # subcontratas
    "subcontrata": "nombre", "nombre subcontrata": "nombre",
    "cif": "cif", "cif / nif": "cif",
    "especialidad": "especialidad",
    "persona de contacto": "contacto", "contacto": "contacto",
    "importe contratado": "importe_contratado", "contratado": "importe_contratado",
    "importe contratado (e)": "importe_contratado",
    "importe certificado": "importe_certificado", "certificado": "importe_certificado",
    "importe pagado": "importe_pagado", "pagado": "importe_pagado",
    "seguro rc": "seguro_rc", "seguro rc vigente": "seguro_rc",
    "prl ok": "prl_ok", "prl / coord. entregada": "prl_ok",
    # maquinaria
    "equipo": "equipo", "equipo / maquina": "equipo",
    "propiedad": "propiedad", "matricula": "matricula",
    "matricula / n serie": "matricula", "matricula / n de serie": "matricula",
    "operario": "operario", "operario asignado": "operario",
    "fecha entrada obra": "fecha_entrada", "entrada": "fecha_entrada",
    "fecha salida obra": "fecha_salida", "salida": "fecha_salida",
    "coste / dia": "coste_dia", "coste/dia": "coste_dia", "coste / dia (e)": "coste_dia",
    "proxima itv": "proxima_itv", "proxima itv / revision": "proxima_itv",
    # incidencias
    "gravedad": "gravedad", "descripcion": "descripcion",
    "zona / ubicacion": "zona", "detectada por": "detectada_por",
    "fecha limite": "fecha_limite", "accion correctiva": "accion_correctiva",
    "coste asociado": "coste", "cerrada el": "fecha_cierre",
    # documentos / planos
    "documento": "documento", "documentacion": "documento",
    "version": "version", "archivo": "enlace", "archivo / enlace": "enlace",
    "fecha de entrega": "fecha_entrega", "entregado el": "fecha_entrega",
    "nombre del plano": "nombre", "disciplina": "disciplina",
    "revision": "revision", "revision actual": "revision",
    "fecha revision": "fecha_revision", "fecha de la revision": "fecha_revision",
    "motivo del cambio": "motivo_cambio",
    "motivo del cambio / edicion": "motivo_cambio",
    # ofertas
    "descripcion de la ampliacion": "descripcion", "motivo": "motivo",
    "importe": "importe", "importe ofertado": "importe", "importe ofertado (e)": "importe",
    "fecha de envio": "fecha_envio", "enviada": "fecha_envio",
    "fecha respuesta cliente": "fecha_respuesta", "respuesta": "fecha_respuesta",
    # presupuesto
    "concepto": "concepto", "presupuestado": "presupuestado",
    "presupuestado (e)": "presupuestado",
    "real": "real", "real (e)": "real", "real / comprometido": "real",
}

# Sinónimos que sólo valen para una entidad concreta, porque la misma cabecera
# significa cosas distintas según la hoja ("Cantidad" es stock en Materiales y
# unidades instaladas en Dispositivos).
SINONIMOS_ENTIDAD: dict[str, dict[str, str]] = {
    "materiales": {
        "cantidad": "recibido", "cantidad recibida": "recibido", "recibida": "recibido",
        "stock": "recibido", "unidades": "recibido",
    },
    "dispositivos": {
        "cantidad": "cantidad", "nombre": "etiqueta", "descripcion": "etiqueta",
        "dispositivo": "etiqueta", "equipo": "etiqueta",
    },
    "tareas": {"nombre": "tarea", "hito": "tarea", "descripcion": "tarea"},
    "personal": {"contratado": "notas", "tipo": "oficio"},
    "partidas": {"categoria": "categoria", "descripcion": "concepto"},
    "documentos": {"nombre": "documento", "descripcion": "documento"},
    "consumos": {"material": "_material_nombre"},
}

VALORES_EJEMPLO = {"ejemplo", "xxx", "-", "n/a", "na", "texto", "…"}

# Cabeceras del tipo "Material 3" / "Cantidad 2": el libro original de Google
# Sheets apunta el material consumido en la propia fila de la tarea.
RE_MATERIAL_N = re.compile(r"^material\s*(\d*)$")
RE_CANTIDAD_N = re.compile(r"^cantidad\s*(\d*)$")


def _norm(texto: Any) -> str:
    s = str(texto or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    s = s.replace("º", "").replace("(€)", "(e)").replace("€", "e")
    s = re.sub(r"\s+", " ", s)
    return s.strip(" :·—-")


def _mapa_campos(entidad: str) -> dict[str, str]:
    """cabecera normalizada -> nombre de campo."""
    ent = ENTITIES[entidad]
    mapa: dict[str, str] = {}
    for f in ent.fields:
        mapa[_norm(f.label)] = f.name
        mapa[_norm(f.name)] = f.name
        mapa[_norm(f.name.replace("_", " "))] = f.name
    validos = {f.name for f in ent.fields}
    for cab, campo in SINONIMOS.items():
        if campo in validos or campo.startswith("_"):
            mapa.setdefault(_norm(cab), campo)
    # Los específicos de la entidad mandan sobre los generales.
    for cab, campo in SINONIMOS_ENTIDAD.get(entidad, {}).items():
        if campo in validos or campo.startswith("_"):
            mapa[_norm(cab)] = campo
    return mapa


def _fila_cabecera(ws) -> int | None:
    # En modo read_only, una hoja vacía devuelve max_row/max_column a None.
    filas, columnas = ws.max_row or 0, ws.max_column or 0
    if not filas or not columnas:
        return None
    mejor, puntos = None, 0
    for r in range(1, min(9, filas + 1)):
        n = 0
        for c in range(1, min(columnas, 45) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and 0 < len(v.strip()) < 45:
                n += 1
        if n > puntos:
            mejor, puntos = r, n
    return mejor if puntos >= 3 else None


def _convertir(f, valor):
    if valor is None:
        return None
    if isinstance(valor, str):
        valor = valor.strip()
        if not valor or _norm(valor) in VALORES_EJEMPLO:
            return None
        if valor.startswith("="):
            return None
    if f.type in ("date", "datetime"):
        if isinstance(valor, (datetime, date)):
            return valor.strftime("%Y-%m-%d")
        try:
            return datetime.fromisoformat(str(valor)[:10]).strftime("%Y-%m-%d")
        except ValueError:
            for formato in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
                try:
                    return datetime.strptime(str(valor).strip(), formato).strftime("%Y-%m-%d")
                except ValueError:
                    continue
            return None
    if f.type in ("number", "money", "percent", "int", "ref"):
        try:
            return float(str(valor).replace("€", "").replace("%", "").replace(",", ".").strip())
        except (TypeError, ValueError):
            return None
    if f.type == "bool":
        return 1 if _norm(valor) in ("si", "sí", "true", "1", "x", "verdadero") else 0
    return str(valor)


def analizar(ruta: Path) -> dict:
    """Inspecciona el libro y dice qué se importaría, sin tocar la base de datos."""
    import openpyxl
    # Sin read_only: los libros exportados de Google Sheets no traen los
    # metadatos de dimensión y en ese modo max_row/max_column salen None.
    wb = openpyxl.load_workbook(ruta, data_only=True)
    informe = []
    for ws in wb.worksheets:
        destino = HOJAS.get(_norm(ws.title))
        if not destino:
            informe.append({"hoja": ws.title, "destino": None, "filas": 0,
                            "columnas": [], "motivo": "Hoja no reconocida"})
            continue
        hr = _fila_cabecera(ws)
        if not hr:
            informe.append({"hoja": ws.title, "destino": destino, "filas": 0,
                            "columnas": [], "motivo": "No se encontró la cabecera"})
            continue
        mapa = _mapa_campos(destino)
        cols, sin_mapear = [], []
        for c in range(1, min(ws.max_column, 45) + 1):
            cab = ws.cell(row=hr, column=c).value
            if not cab:
                continue
            campo = mapa.get(_norm(cab))
            (cols if campo else sin_mapear).append(str(cab))
        filas = max(0, (ws.max_row or 0) - hr)
        informe.append({
            "hoja": ws.title, "destino": destino, "entidad": ENTITIES[destino].plural,
            "filas": filas, "columnas": cols, "ignoradas": sin_mapear,
        })
    wb.close()
    return {"hojas": informe}


def importar(ruta: Path, obra_id: int, usuario: str | None = None,
             solo_hojas: list[str] | None = None) -> dict:
    """Importa el libro a la obra indicada. Devuelve el recuento por entidad."""
    import openpyxl
    # Sin read_only: los libros exportados de Google Sheets no traen los
    # metadatos de dimensión y en ese modo max_row/max_column salen None.
    wb = openpyxl.load_workbook(ruta, data_only=True)
    resultado: dict[str, int] = {}
    avisos: list[str] = []
    pendientes: list[dict] = []   # consumos por resolver al terminar

    # Para resolver las referencias de consumos -> tareas / materiales por nombre.
    idx_tareas = {_norm(t["tarea"]): t["id"] for t in db.listar("tareas", obra_id) if t.get("tarea")}
    idx_materiales = {_norm(m["material"]): m["id"] for m in db.listar("materiales", obra_id) if m.get("material")}

    for ws in wb.worksheets:
        destino = HOJAS.get(_norm(ws.title))
        if not destino or (solo_hojas and ws.title not in solo_hojas):
            continue
        hr = _fila_cabecera(ws)
        if not hr:
            continue
        ent = ENTITIES[destino]
        campos = {f.name: f for f in ent.fields}
        mapa = _mapa_campos(destino)

        columnas: dict[int, str] = {}
        pares_material: dict[str, dict[str, int]] = {}
        for c in range(1, min(ws.max_column or 0, 60) + 1):
            cab = ws.cell(row=hr, column=c).value
            if not cab:
                continue
            norm = _norm(cab)
            # En la hoja de tareas, las parejas "Material N"/"Cantidad N" son
            # consumos apuntados en la propia fila: se rescatan aparte.
            if destino == "tareas":
                if (m := RE_MATERIAL_N.match(norm)):
                    pares_material.setdefault(m.group(1), {})["material"] = c
                    continue
                if (m := RE_CANTIDAD_N.match(norm)):
                    pares_material.setdefault(m.group(1), {})["cantidad"] = c
                    continue
            if (campo := mapa.get(norm)):
                columnas[c] = campo
        pares_material = {k: v for k, v in pares_material.items()
                          if "material" in v and "cantidad" in v}
        if not columnas:
            continue

        n = 0
        for r in range(hr + 1, (ws.max_row or hr) + 1):
            datos: dict[str, Any] = {}
            extra: dict[str, Any] = {}
            for c, campo in columnas.items():
                bruto = ws.cell(row=r, column=c).value
                if campo.startswith("_"):
                    extra[campo] = bruto
                    continue
                v = _convertir(campos[campo], bruto)
                if v is not None:
                    datos[campo] = v
            if not datos:
                continue
            # Una fila cuenta si tiene relleno el campo que da título a la entidad.
            if not datos.get(ent.title_field):
                continue

            if pares_material:
                extra["_consumos"] = []
                for par in pares_material.values():
                    nombre = ws.cell(row=r, column=par["material"]).value
                    cant = ws.cell(row=r, column=par["cantidad"]).value
                    if not nombre or cant in (None, "", 0):
                        continue
                    try:
                        cantidad = float(str(cant).replace(",", "."))
                    except (TypeError, ValueError):
                        continue
                    extra["_consumos"].append((str(nombre).strip(), cantidad))

            creado = db.crear(destino, datos, obra_id=obra_id, usuario=usuario)
            n += 1

            if destino == "tareas":
                idx_tareas[_norm(datos["tarea"])] = creado["id"]
                for nombre_mat, cantidad in extra.get("_consumos", []):
                    pendientes.append({
                        "tarea_id": creado["id"], "material": nombre_mat,
                        "cantidad": cantidad, "fecha": datos.get("fecha_inicio"),
                    })
            elif destino == "materiales":
                idx_materiales[_norm(datos["material"])] = creado["id"]
            elif destino == "consumos":
                nombre_mat = extra.get("_material_nombre")
                if nombre_mat and not datos.get("material_id"):
                    pendientes.append({
                        "consumo_id": creado["id"], "material": str(nombre_mat),
                    })

        if n:
            resultado[ent.plural] = resultado.get(ent.plural, 0) + n

    wb.close()

    # Los consumos se resuelven al final: la hoja de tareas puede venir antes
    # que la de materiales, y hasta entonces no se sabe a qué material apuntan.
    creados_consumo = 0
    for p in pendientes:
        clave = _norm(p["material"])
        mid = idx_materiales.get(clave)
        if mid is None:
            # El material no estaba en el libro: se crea para no perder el consumo.
            nuevo = db.crear("materiales", {
                "material": p["material"], "recibido": 0,
                "notas": "Creado automáticamente al importar un consumo.",
            }, obra_id=obra_id, usuario=usuario)
            mid = idx_materiales[clave] = nuevo["id"]
            avisos.append(f"Material «{p['material']}» creado a partir de un consumo.")
        if "consumo_id" in p:
            db.actualizar("consumos", p["consumo_id"], {"material_id": mid}, usuario)
        else:
            db.crear("consumos", {
                "material_id": mid, "cantidad": p["cantidad"],
                "tarea_id": p["tarea_id"], "fecha": p.get("fecha"),
                "notas": "Importado desde la fila de la tarea.",
            }, obra_id=obra_id, usuario=usuario)
            creados_consumo += 1
    if creados_consumo:
        resultado["Consumos"] = resultado.get("Consumos", 0) + creados_consumo

    db.registrar(usuario, "importar", None, None, obra_id,
                 f"{ruta.name}: " + ", ".join(f"{k}={v}" for k, v in resultado.items()))
    return {"importado": resultado, "avisos": avisos}


def datos_obra_desde_libro(ruta: Path) -> dict:
    """Intenta leer la ficha del proyecto de la hoja Portada para prerrellenar la obra."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception:
        return {}
    campos = {
        "nombre del proyecto": "nombre", "proyecto": "nombre",
        "cliente": "cliente", "ubicacion": "poblacion", "direccion": "direccion",
        "jefe de obra": "jefe_obra", "telefono": "telefono_jefe", "email": "email_jefe",
        "fecha de inicio": "fecha_inicio", "fecha fin prevista": "fecha_fin_prevista",
        "n oferta / contrato": "num_oferta", "n oferta": "num_oferta",
        "estado general": "estado",
    }
    # La portada de estos libros lleva la ficha a la izquierda y un índice de
    # hojas a la derecha. Si se busca "el primer valor no vacío de la fila" se
    # acaba capturando el índice, así que se descartan los nombres de hoja.
    nombres_hoja = {_norm(h) for h in wb.sheetnames}

    salida: dict[str, Any] = {}
    for ws in wb.worksheets:
        if _norm(ws.title) not in ("portada", "ficha", "inicio", "general"):
            continue
        filas, columnas = ws.max_row or 0, ws.max_column or 0
        for r in range(1, min(filas + 1, 40)):
            for c in range(1, min(columnas + 1, 8)):
                campo = campos.get(_norm(ws.cell(row=r, column=c).value))
                if not campo:
                    continue
                # Sólo la celda contigua (o la siguiente, si aquélla está combinada).
                for cc in (c + 1, c + 2):
                    if cc > columnas:
                        break
                    v = ws.cell(row=r, column=cc).value
                    if v in (None, "") or str(v).startswith("="):
                        continue
                    if isinstance(v, (datetime, date)):
                        salida.setdefault(campo, v.strftime("%Y-%m-%d"))
                        break
                    texto = str(v).strip()
                    if _norm(texto) in nombres_hoja or _norm(texto) in VALORES_EJEMPLO:
                        break  # es una entrada del índice, no un dato de la ficha
                    salida.setdefault(campo, texto)
                    break
    wb.close()
    return salida
